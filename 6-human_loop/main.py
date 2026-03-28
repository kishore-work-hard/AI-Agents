"""
Level 6 — Human in the Loop Agent
───────────────────────────────────
The agent can take REAL actions on your system.
Before any dangerous/irreversible action, it pauses
and asks for your approval.

Safe actions   → execute automatically
Dangerous actions → pause, show plan, wait for yes/no

Real tools available:
  • read_file    — read a file's contents       (SAFE)
  • list_files   — list files in a directory    (SAFE)
  • write_file   — create or overwrite a file   (DANGEROUS)
  • delete_file  — permanently delete a file    (DANGEROUS)
  • run_command  — run a shell command          (DANGEROUS)
"""

import os
import json
import subprocess
import re
import sys
from groq import Groq

# ── Setup ─────────────────────────────────────────────────────────────────────

client = Groq(api_key="")
MODEL = "moonshotai/kimi-k2-instruct"


# ── Tool definitions ──────────────────────────────────────────────────────────
# Each tool has:
#   fn        — the actual Python function to call
#   dangerous — True = needs human approval before running
#   describe  — returns a human-readable description of what it WILL do
#               (shown to user BEFORE approval)

def _read_file(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        return content if content else "(empty file)"
    except FileNotFoundError:
        return f"Error: file not found — {path}"
    except Exception as e:
        return f"Error reading file: {e}"


def _list_files(path: str = ".") -> str:
    try:
        entries = os.listdir(path)
        if not entries:
            return "(empty directory)"
        files = []
        for e in sorted(entries):
            full = os.path.join(path, e)
            tag  = "[DIR]" if os.path.isdir(full) else "[FILE]"
            files.append(f"  {tag} {e}")
        return "\n".join(files)
    except Exception as e:
        return f"Error listing directory: {e}"


def _write_file(path: str, content: str) -> str:
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"Successfully wrote to {path}"
    except Exception as e:
        return f"Error writing file: {e}"


def _delete_file(path: str) -> str:
    try:
        os.remove(path)
        return f"Successfully deleted {path}"
    except FileNotFoundError:
        return f"Error: file not found — {path}"
    except Exception as e:
        return f"Error deleting file: {e}"


def _run_command(command: str) -> str:
    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=30
        )
        output = result.stdout.strip()
        errors = result.stderr.strip()
        if output and errors:
            return f"Output:\n{output}\n\nErrors:\n{errors}"
        return output or errors or "(no output)"
    except subprocess.TimeoutExpired:
        return "Error: command timed out after 30 seconds"
    except Exception as e:
        return f"Error running command: {e}"


def _read_excel(path: str) -> str:
    """
    Read an Excel file (.xlsx/.xls) and return its contents as a
    readable text table. Uses openpyxl which is already installed
    in the venv — no pip install needed.
    """
    try:
        import openpyxl
    except ImportError:
        # Try installing into the CURRENT venv automatically
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            capture_output=True
        )
        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl not available. Run: pip install openpyxl"

    try:
        wb    = openpyxl.load_workbook(path, data_only=True)
        output = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            output.append(f"\n=== Sheet: {sheet_name} ===")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                output.append("(empty sheet)")
                continue

            # Find max columns that actually have data
            max_col = max(
                (len([c for c in row if c is not None]) for row in rows),
                default=0
            )
            if max_col == 0:
                output.append("(empty sheet)")
                continue

            # Format as table
            for row in rows:
                cells = [str(c) if c is not None else "" for c in row[:max_col]]
                output.append(" | ".join(cells))

        return "\n".join(output)

    except FileNotFoundError:
        return f"Error: file not found — {path}"
    except Exception as e:
        return f"Error reading Excel file: {e}"


# Tool registry — name → config
TOOLS = {
    "read_file": {
        "fn":        _read_file,
        "dangerous": False,
        "describe":  lambda args: f"Read file: {args.get('path', '?')}",
    },
    "list_files": {
        "fn":        _list_files,
        "dangerous": False,
        "describe":  lambda args: f"List files in: {args.get('path', '.')}",
    },
    "write_file": {
        "fn":        _write_file,
        "dangerous": True,
        "describe":  lambda args: (
            f"Write to file: {args.get('path', '?')}\n"
            f"  Content preview: {str(args.get('content', ''))[:100]}..."
            if len(str(args.get('content', ''))) > 100
            else f"Write to file: {args.get('path', '?')}\n  Content: {args.get('content', '')}"
        ),
    },
    "delete_file": {
        "fn":        _delete_file,
        "dangerous": True,
        "describe":  lambda args: f"PERMANENTLY DELETE file: {args.get('path', '?')}",
    },
    "run_command": {
        "fn":        _run_command,
        "dangerous": True,
        "describe":  lambda args: f"Run shell command: {args.get('command', '?')}",
    },
    "read_excel": {
        "fn":        _read_excel,
        "dangerous": False,
        "describe":  lambda args: f"Read Excel file: {args.get('path', '?')}",
    },
}


# ── System prompt ─────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a helpful file system assistant that can read, write, and manage files.

AVAILABLE TOOLS:

1. read_file(path)
   - Read contents of a file
   - Safe — no approval needed
   - Args: {"path": "filename.txt"}

2. list_files(path)
   - List all files in a directory
   - Safe — no approval needed
   - Args: {"path": "."} (use "." for current directory)

3. write_file(path, content)
   - Create or overwrite a file with content
   - DANGEROUS — requires human approval
   - Args: {"path": "filename.txt", "content": "file content here"}

4. delete_file(path)
   - Permanently delete a file
   - DANGEROUS — requires human approval
   - Args: {"path": "filename.txt"}

5. run_command(command)
   - Run a shell command
   - DANGEROUS — requires human approval
   - Args: {"command": "echo hello"}

6. read_excel(path)
   - Read an Excel file (.xlsx or .xls) and show its contents as a table
   - Safe — no approval needed
   - Args: {"path": "C:\\Users\\KISHORE\\Desktop\\file.xlsx"}
   - Use this for ANY .xlsx or .xls file instead of read_file

RESPONSE FORMAT — always use one of these two:

Format 1 — call a tool:
THOUGHT: <your reasoning>
ACTION: <tool name exactly as listed>
ARGS: <valid JSON with the arguments>

Format 2 — final answer (when done):
THOUGHT: <your reasoning>
FINAL ANSWER: <your response to the user>

RULES:
- Always use ARGS as valid JSON
- One tool call per response
- After getting an OBSERVATION, decide if you need more tools or can give FINAL ANSWER
- Never guess file contents — always read them first
- Never run destructive commands without clear user intent
- For .xlsx or .xls files ALWAYS use read_excel, never read_file
- NEVER run pip install — use the tools provided instead"""


# ── Response parser ───────────────────────────────────────────────────────────

def parse_response(text: str) -> dict:
    """
    Parse AI response into structured dict.
    Returns one of:
        {"type": "action",  "action": "...", "args": {...}, "thought": "..."}
        {"type": "final",   "answer": "...", "thought": "..."}
        {"type": "unknown", "raw": "..."}
    """
    result = {}
    lines  = text.strip().split("\n")

    # Collect multi-line ARGS
    collecting_args = False
    args_lines      = []

    for line in lines:
        stripped = line.strip()

        if stripped.startswith("THOUGHT:"):
            result["thought"] = stripped.replace("THOUGHT:", "").strip()
            collecting_args   = False

        elif stripped.startswith("ACTION:"):
            result["type"]    = "action"
            result["action"]  = stripped.replace("ACTION:", "").strip()
            collecting_args   = False

        elif stripped.startswith("ARGS:"):
            collecting_args = True
            inline = stripped.replace("ARGS:", "").strip()
            if inline:
                args_lines = [inline]
            else:
                args_lines = []

        elif stripped.startswith("FINAL ANSWER:"):
            result["type"]   = "final"
            result["answer"] = stripped.replace("FINAL ANSWER:", "").strip()
            collecting_args  = False

        elif collecting_args:
            args_lines.append(stripped)

    # Parse collected ARGS JSON
    if args_lines:
        raw_args = " ".join(args_lines)
        # Extract JSON object from the string
        json_match = re.search(r'\{.*\}', raw_args, re.DOTALL)
        if json_match:
            try:
                result["args"] = json.loads(json_match.group())
            except json.JSONDecodeError:
                result["args"] = {}
        else:
            result["args"] = {}

    if "type" not in result:
        result["type"] = "unknown"
        result["raw"]  = text

    return result


# ── Human approval gate ───────────────────────────────────────────────────────

def request_approval(tool_name: str, args: dict, describe_fn) -> bool:
    """
    Show the user exactly what the agent wants to do.
    Return True if approved, False if rejected.
    This is the core of "Human in the Loop" —
    dangerous actions CANNOT bypass this gate.
    """
    description = describe_fn(args)

    print(f"\n{'⚠️ ' * 20}")
    print(f"  APPROVAL REQUIRED")
    print(f"{'─' * 40}")
    print(f"  Tool   : {tool_name}")
    print(f"  Action : {description}")
    print(f"{'─' * 40}")

    while True:
        answer = input("  Approve? (yes/no): ").strip().lower()
        if answer in ("yes", "y"):
            print(f"  ✅ Approved")
            return True
        elif answer in ("no", "n"):
            print(f"  ❌ Rejected")
            return False
        else:
            print("  Please type 'yes' or 'no'")


# ── Tool executor ─────────────────────────────────────────────────────────────

def execute_tool(tool_name: str, args: dict) -> str:
    """
    Execute a tool with human approval gate for dangerous tools.
    Returns the tool's output as a string (the observation).
    """
    if tool_name not in TOOLS:
        return f"Error: unknown tool '{tool_name}'"

    tool = TOOLS[tool_name]

    # ── Dangerous tool: ask for approval first ────────────────────────────────
    if tool["dangerous"]:
        approved = request_approval(tool_name, args, tool["describe"])
        if not approved:
            return f"Action cancelled by user."

    # ── Execute the tool ──────────────────────────────────────────────────────
    try:
        fn = tool["fn"]

        # Call the function with the right arguments
        # Each tool has different parameter names so we unpack from args dict
        if tool_name == "read_file":
            return fn(args.get("path", ""))

        elif tool_name == "list_files":
            return fn(args.get("path", "."))

        elif tool_name == "write_file":
            return fn(args.get("path", ""), args.get("content", ""))

        elif tool_name == "delete_file":
            return fn(args.get("path", ""))

        elif tool_name == "run_command":
            return fn(args.get("command", ""))

        elif tool_name == "read_excel":
            return fn(args.get("path", ""))

        else:
            return f"Error: no handler for tool '{tool_name}'"

    except Exception as e:
        return f"Tool execution error: {e}"


# ── ReAct loop (same pattern as Level 4 but with real tools + approval) ───────

def run_agent(user_question: str, messages: list, max_steps: int = 10) -> str:
    """
    Run the agent loop:
    THINK → ACT (with approval if dangerous) → OBSERVE → repeat → FINAL ANSWER

    messages is passed in from outside so conversation history persists
    across multiple user questions — the agent remembers everything.
    """
    print(f"\n{'═' * 60}")
    print(f"Task: {user_question}")
    print(f"{'═' * 60}")

    # Add the new user question to the existing conversation history
    # This is the key change — we don't reset messages, we append to it
    messages.append({"role": "user", "content": user_question})

    for step in range(max_steps):
        print(f"\n--- Step {step + 1} ---")

        # Ask AI what to do next
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0,
            max_tokens=600,
        )

        ai_text = response.choices[0].message.content
        print(f"\nAI thinking:\n{ai_text}")

        parsed = parse_response(ai_text)

        # ── Agent wants to use a tool ─────────────────────────────────────────
        if parsed["type"] == "action":
            tool_name = parsed.get("action", "")
            args      = parsed.get("args", {})

            # Add AI message to history
            messages.append({"role": "assistant", "content": ai_text})

            # Execute tool (with approval gate if dangerous)
            print(f"\n→ Tool: {tool_name} | Args: {args}")
            observation = execute_tool(tool_name, args)
            print(f"← Result: {observation[:200]}{'...' if len(observation) > 200 else ''}")

            # For write/delete — if successful, return immediately.
            # The model sometimes replies with just "File deleted." after
            # these ops which breaks the parser and causes a nudge loop.
            # The action already succeeded so there's nothing left to do.
            if tool_name in ("write_file", "delete_file") and "Successfully" in observation:
                messages.append({"role": "assistant", "content": observation})
                print(f"\n✅ Done: {observation}")
                return observation

            # Add observation back into conversation
            messages.append({
                "role":    "user",
                "content": f"OBSERVATION: {observation}"
            })

        # ── Agent has final answer ────────────────────────────────────────────
        elif parsed["type"] == "final":
            answer = parsed.get("answer", "")
            # Add final answer to history so next question has context
            messages.append({"role": "assistant", "content": answer})
            print(f"\n✅ Done: {answer}")
            return answer

        # ── Could not parse ───────────────────────────────────────────────────
        else:
            print("⚠️  Could not parse response, nudging agent...")
            messages.append({"role": "assistant", "content": ai_text})
            messages.append({
                "role":    "user",
                "content": "Please respond using exactly: THOUGHT, ACTION, ARGS or THOUGHT, FINAL ANSWER"
            })

    return "Task incomplete — reached maximum steps."


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("╔══════════════════════════════════════╗")
    print("║   Level 6 — Human in the Loop Agent  ║")
    print("╠══════════════════════════════════════╣")
    print("║  Safe actions    → auto execute       ║")
    print("║  Dangerous actions → asks you first   ║")
    print("╠══════════════════════════════════════╣")
    print("║  Tools available:                     ║")
    print("║    read_file    ✅ safe               ║")
    print("║    read_excel   ✅ safe               ║")
    print("║    list_files   ✅ safe               ║")
    print("║    write_file   ⚠️  approval needed   ║")
    print("║    delete_file  ⚠️  approval needed   ║")
    print("║    run_command  ⚠️  approval needed   ║")
    print("╚══════════════════════════════════════╝")
    print("\nType 'quit' to exit\n")

    # Messages live here — persists for the whole session
    # System prompt is added once at the start
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT}
    ]

    while True:
        user_input = input("You: ").strip()
        if not user_input:
            continue
        if user_input.lower() in ("quit", "exit"):
            break

        result = run_agent(user_input, messages)
        print(f"\nAgent: {result}\n")


if __name__ == "__main__":
    main()
